import torch
import numpy as np
import pandas as pd
import torchvision
from d2l.torch import get_dataloader_workers
from torchvision import transforms
from torch import nn
from d2l import torch as d2l
from torch.utils import data
from train_network.calculate_tool import data_preprocess1
from train_network.calculate_tool.data_preprocess1 import timefn
from openpyxl import load_workbook


# the function that generates the dataset which is for a single Excel sheet
def generate_dataset3(file_path, sheet_name, labels_preprocess=True, return_current_board=False):
    data = pd.read_excel(file_path, sheet_name)

    np.set_printoptions(precision=0, threshold=np.inf, linewidth=np.inf)
    torch.set_printoptions(precision=4, sci_mode=False)
    np.set_printoptions(precision=4)

    # read flood data
    in_flow = data.iloc[0:, 0]
    out_flow = data.iloc[0:, 1]
    rain = data.iloc[0:, 2]
    z_up = data.iloc[0:, 3]
    z_down = data.iloc[0:, 4]
    in_flow_rise = data.iloc[0:, 5]
    in_flow_reduce = data.iloc[0:, 6]
    out_flow_rise = data.iloc[0:, 7]
    out_flow_reduce = data.iloc[0:, 8]
    out_flow_extent = data.iloc[0:, 9]
    out_flow_limit = data.iloc[0:, 11]
    water_gate = data.iloc[0:, 12]
    reservoir_feature = data.iloc[0:20, 14]
    flood_evaluate = data.iloc[20, 16]

    # Initializes the data matrix
    flood_matrix = np.zeros([24, 60, 60], dtype=np.float16)

    # Create a four-dimensional array to store the data matrix
    features_matrix = np.zeros([len(out_flow), 24, 60, 60], dtype=np.float16)
    # Create a four-dimensional array to store the label matrix
    labels_matrix1 = np.zeros(len(out_flow), dtype=int)
    labels_matrix2 = np.zeros(len(out_flow), dtype=np.float16)

    out_flow_num = len([flow for flow in out_flow if not np.isnan(flow)])
    if out_flow_num == len(in_flow):
        pass
    else:
        out_flow_num = out_flow_num + 1

    for j in range(out_flow_num):
        a = data_preprocess1.format_data4_flow(in_flow[j])
        c = data_preprocess1.format_data4_rain(rain[j])
        if not j == 0:
            b = data_preprocess1.format_data4_flow(out_flow[j - 1])
            d = data_preprocess1.format_data5(z_up[j - 1])
            e = data_preprocess1.format_data5(z_down[j - 1])
            f = data_preprocess1.format_data4_flow(in_flow_rise[j - 1])
            h = data_preprocess1.format_data4_flow(in_flow_reduce[j - 1])
            o = data_preprocess1.format_data4_flow(out_flow_rise[j - 1])
            p = data_preprocess1.format_data4_flow(out_flow_reduce[j - 1])

        m = int((j - 1) // (flood_matrix.shape[2] / 2))
        n = int((j - 1) % (flood_matrix.shape[2] / 2))
        m_1 = int(j // (flood_matrix.shape[2] / 2))
        n_1 = int(j % (flood_matrix.shape[2] / 2))

        # feature matrix filling
        if j == 0:
            count = 0
            count = count + 1

            if not np.isnan(a).any():
                for i in range(len(a)):
                    flood_matrix[i + count, a[i], 0] = 1

        else:
            # fill flood frequency（first layer）
            count = 0
            q = data_preprocess1.format_data7(reservoir_feature[0])
            q = q[3:]
            if not np.isnan(q).any():
                for i in range(features_matrix.shape[2] // 10):
                    for k in range(len(q)):
                        flood_matrix[count, i * 10:(i + 1) * 10, q[k] + k * 10] = 1
            count = count + 1

            # fill inflow flood and outflow values（2nd-5th layer）
            if not np.isnan(a).any():
                for i in range(len(a)):
                    flood_matrix[i + count, a[i] + m_1 * 10, n_1 * 2] = 1
            if not np.isnan(b).any():
                for i in range(len(b)):
                    flood_matrix[i + count, b[i] + m * 10, n * 2 + 1] = 1
            count = count + 4

            # fill all 1 matrix（6th layer）
            flood_matrix[count, :, :] = 1
            count = count + 1

            # fill water levels above and below the dam（7th-11th layer）
            if not np.isnan(d).any():
                for i in range(len(d)):
                    flood_matrix[i + count, d[i] + m * 10, n * 2] = 1
            if not np.isnan(e).any():
                for i in range(len(e)):
                    flood_matrix[i + count, e[i] + m * 10, n * 2 + 1] = 1
            count = count + 5

            # fill rates of inflow increase and decrease（12th-15th layer）
            if not np.isnan(f).any():
                for i in range(len(f)):
                    flood_matrix[i + count, f[i] + m * 10, n * 2] = 1
            if not np.isnan(h).any():
                for i in range(len(h)):
                    flood_matrix[i + count, h[i] + m * 10, n * 2 + 1] = 1
            count = count + 4

            # fill rates of outflow increase and decrease（16th-19th layer）
            if not np.isnan(o).any():
                for i in range(len(o)):
                    flood_matrix[i + count, o[i] + m * 10, n * 2] = 1
            if not np.isnan(p).any():
                for i in range(len(p)):
                    flood_matrix[i + count, p[i] + m * 10, n * 2 + 1] = 1
            count = count + 4

            # fill flood peak time（20th layer）
            if not np.isnan(reservoir_feature[len(reservoir_feature) - 1]):
                flood_matrix[count, :, :] = reservoir_feature[len(reservoir_feature) - 1]
            count = count + 1

            # fill flood shape（21th layer）
            if not np.isnan(reservoir_feature[len(reservoir_feature) - 2]):
                flood_matrix[count, :, :] = reservoir_feature[len(reservoir_feature) - 2]
            count = count + 1

            # fill limitation of inflow variation（22th layer）
            if not np.isnan(out_flow_extent[j - 1]):
                flood_matrix[count, m * 10: (m + 1) * 10, (n * 2):(n + 1) * 2] = out_flow_extent[j - 1]
            count = count + 1

            # fill limitation of outflow variation（23th layer）
            if not np.isnan(out_flow_limit[j - 1]):
                flood_matrix[count, m * 10: (m + 1) * 10, (n * 2):(n + 1) * 2] = out_flow_limit[j - 1]
            count = count + 1

            # fill all 0 matrix（24th layer）
            flood_matrix[count, :, :] = 0

        # Policy network label
        if labels_preprocess:
            if not np.isnan(out_flow[j]).any():
                labels_location = round(out_flow[j] / 50) - 1
                if labels_location == -1:
                    labels_location = 0
                labels_matrix1[j] = labels_location
        else:
            if not np.isnan(out_flow[j]).any():
                labels_location = out_flow[j]
                labels_matrix1[j] = labels_location

        # value network label
        labels_matrix2[j] = round(flood_evaluate, 4)

        # Store data matrix
        features_matrix[j] = flood_matrix

    if return_current_board:
        return features_matrix[out_flow_num - 1], labels_matrix1[out_flow_num - 1]
    else:
        if labels_preprocess:
            return features_matrix, labels_matrix1, labels_matrix2
        else:
            return features_matrix, labels_matrix1, labels_matrix2, in_flow, z_up, z_down


# the function that generates the dataset which is for all Excel sheets
@timefn
def generate_dataset_allexcel3(file_path, labels_preprocess=True):
    global features_matrix, policy_labels_matrix, value_labels_matrix, in_flow, z_up, z_down
    sheets_dict = pd.read_excel(file_path, sheet_name=None)
    if labels_preprocess:
        sheet_1_features, sheet_1_policy_labels, sheet_1_value_labels = generate_dataset3(
            file_path, sheet_name=1, labels_preprocess=labels_preprocess)
    else:
        sheet_1_features, sheet_1_policy_labels, sheet_1_value_labels, in_flow, z_up, z_down = generate_dataset3(
            file_path, sheet_name=1, labels_preprocess=labels_preprocess)
    for i in range(len(sheets_dict)):
        # Skip the first sheet
        if i == 0:
            pass
        if i == 1:
            features_matrix = sheet_1_features
            policy_labels_matrix = sheet_1_policy_labels
            value_labels_matrix = sheet_1_value_labels

        if i > 1:
            # Join multiple four-dimensional matrices
            if labels_preprocess:
                sheet_j_features, sheet_j_policy_labels, sheet_j_value_labels = generate_dataset3(
                    file_path, sheet_name=i, labels_preprocess=labels_preprocess)
            else:
                sheet_j_features, sheet_j_policy_labels, sheet_j_value_labels, _, _, _ = generate_dataset3(
                    file_path, sheet_name=i, labels_preprocess=labels_preprocess)

            features_matrix = np.concatenate((features_matrix, sheet_j_features), axis=0)
            policy_labels_matrix = np.concatenate((policy_labels_matrix, sheet_j_policy_labels), axis=0)
            value_labels_matrix = np.concatenate((value_labels_matrix, sheet_j_value_labels), axis=0)

    if labels_preprocess:
        print(f"features_matrix.shape={features_matrix.shape}")
        return features_matrix, policy_labels_matrix, value_labels_matrix
    else:
        return features_matrix, policy_labels_matrix, value_labels_matrix, in_flow, z_up, z_down


@timefn
def generate_dataset_allexcel3_beta(file_path, test_sheet_names, labels_preprocess=True):
    global features_matrix, policy_labels_matrix, value_labels_matrix, in_flow, z_up, z_down
    # Remove sheets from the test dataset
    # Loading Excel file
    workbook = load_workbook(file_path)
    # Gets the names of all worksheets
    sheet_names = workbook.sheetnames
    # Gets indexes for all worksheets
    sheet_indexes = list(range(len(workbook.sheetnames)))
    # Gets the index value based on the worksheet name
    excluded_indexes = []
    for excluded_sheet in test_sheet_names:
        excluded_indexes.append(sheet_names.index(excluded_sheet))
    # Removes the specified index value
    filtered_indexes = [index for index in sheet_indexes if index not in excluded_indexes]

    for i, sheet in enumerate(filtered_indexes):
        # Skip the first sheet
        if i == 0:
            pass
        if i == 1:
            if labels_preprocess:
                features_matrix, policy_labels_matrix, value_labels_matrix = generate_dataset3(
                    file_path, sheet_name=sheet, labels_preprocess=labels_preprocess)
            else:
                features_matrix, policy_labels_matrix, value_labels_matrix, in_flow, z_up, z_down = generate_dataset3(
                    file_path, sheet_name=sheet, labels_preprocess=labels_preprocess)
        if i > 1:
            # Join multiple four-dimensional matrices
            if labels_preprocess:
                sheet_j_features, sheet_j_policy_labels, sheet_j_value_labels = generate_dataset3(
                    file_path, sheet_name=sheet, labels_preprocess=labels_preprocess)
            else:
                sheet_j_features, sheet_j_policy_labels, sheet_j_value_labels, _, _, _ = generate_dataset3(
                    file_path, sheet_name=sheet, labels_preprocess=labels_preprocess)

            features_matrix = np.concatenate((features_matrix, sheet_j_features), axis=0)
            policy_labels_matrix = np.concatenate((policy_labels_matrix, sheet_j_policy_labels), axis=0)
            value_labels_matrix = np.concatenate((value_labels_matrix, sheet_j_value_labels), axis=0)
    if labels_preprocess:
        print(f"features_matrix.shape={features_matrix.shape}")
        return features_matrix, policy_labels_matrix, value_labels_matrix
    else:
        return features_matrix, policy_labels_matrix, value_labels_matrix, in_flow, z_up, z_down
